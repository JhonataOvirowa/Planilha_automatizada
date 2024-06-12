from openpyxl import load_workbook

wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"]

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row